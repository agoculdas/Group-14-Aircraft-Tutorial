"""
Data loader for CRJ-1000 from Adsee3Plane.xlsx
Parses the spreadsheet into clean Python dictionaries.

Sheet layout:
  Part I  — geometry, MAC calculation, aero parameters, scissor-plot inputs
  Part II — updated component CG (incl. batteries), loading diagram, CG extremes
"""

import openpyxl


def load_data(filepath="Adsee3Plane.xlsx"):
    """Load all data from the Excel file into a structured dictionary."""
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Handle both old ("template") and new ("Part I"/"Part II") sheet names
    if "Part I" in wb.sheetnames:
        ws = wb["Part I"]
        ws2 = wb["Part II"] if "Part II" in wb.sheetnames else ws
    else:
        ws = wb["template"]
        ws2 = ws

    def cell(ref, sheet=None):
        return (sheet or ws)[ref].value

    data = {}

    # =========================================================================
    # 1. TOP-LEVEL WEIGHTS
    # =========================================================================
    data["weights"] = {
        "MTOW":          cell("C3"),
        "EOW":           cell("C4"),
        "MMAXPAY":       cell("C5"),
        "FUEL_MAXPAY":   cell("C6"),
        "Cabin_Payload": cell("C7"),
        "Fore_Payload":  cell("C8"),
        "Aft_Payload":   cell("C9"),
    }

    # =========================================================================
    # 2. CG POSITIONS (datum = aircraft nose)
    # =========================================================================
    data["cg_positions_m"] = {
        "EOW":          cell("D4"),
        "FUEL_MAXPAY":  cell("D6"),
        "Fore_Payload": cell("D8"),
        "Aft_Payload":  cell("D9"),
    }

    # =========================================================================
    # 3. ENGINE
    # =========================================================================
    data["engine"] = {
        "type":      cell("G3"),
        "mass_kg":   cell("H3"),
        "thrust_N":  cell("H2"),
        "range_km":  cell("H1"),
    }

    # =========================================================================
    # 4. MAIN WING GEOMETRY
    # =========================================================================
    data["wing"] = {
        "b":     cell("H6"),
        "S":     cell("H7"),
        "AR":    cell("K97")
    }

    # =========================================================================
    # 5. HORIZONTAL TAIL GEOMETRY
    # =========================================================================
    data["HT"] = {
        "b_h":       cell("H10"),
        "c_h_mac":   cell("H11"),
        "c_h_root":  cell("H12"),
        "c_h_tip":   cell("H13"),
        "x_total":   cell("H14"),
    }
    ht = data["HT"]
    ht["taper_ratio"] = ht["c_h_tip"] / ht["c_h_root"]
    ht["S_h"] = 0.5 * ht["b_h"] * (ht["c_h_root"] + ht["c_h_tip"])
    ht["AR_h"] = ht["b_h"]**2 / ht["S_h"]

    # =========================================================================
    # 6. VERTICAL TAIL GEOMETRY
    # =========================================================================
    data["VT"] = {
        "c_v_mac":   cell("H18"),
        "c_v_root":  cell("H19"),
        "c_v_tip":   cell("H20"),
    }

    # =========================================================================
    # 7. AIRCRAFT EXTERNAL DIMENSIONS
    # =========================================================================
    data["dimensions"] = {
        "length":     cell("H24"),
        "height":     cell("H25"),
        "fus_length": cell("H28"),
    }

    # =========================================================================
    # 8. CABIN
    # =========================================================================
    data["cabin"] = {
        "config":      cell("H30"),
        "ext_dim_m":   cell("H31"),
        "pax_num":     cell("H33"),
        "pax_weight":  cell("H34"),
        "seat_pitch":  cell("H36"),
        "first_seat":  cell("H37"),
    }

    # =========================================================================
    # 9. COMPONENT CG BREAKDOWN — from Part II if available (includes batteries)
    # =========================================================================
    data["component_cg"] = {
        "Wing":             {"mass_kg": cell("N4",  ws2), "cg_m": cell("O4",  ws2), "cg_pctMAC": cell("P4",  ws2)},
        "MainLG":           {"mass_kg": cell("N5",  ws2), "cg_m": cell("O5",  ws2), "cg_pctMAC": cell("P5",  ws2)},
        "HorizontalTail":   {"mass_kg": cell("N7",  ws2), "cg_m": cell("O7",  ws2), "cg_pctMAC": cell("P7",  ws2)},
        "VerticalTail":     {"mass_kg": cell("N8",  ws2), "cg_m": cell("O8",  ws2), "cg_pctMAC": cell("P8",  ws2)},
        "Fuselage":         {"mass_kg": cell("N9",  ws2), "cg_m": cell("O9",  ws2), "cg_pctMAC": cell("P9",  ws2)},
        "NoseLG":           {"mass_kg": cell("N10", ws2), "cg_m": cell("O10", ws2), "cg_pctMAC": cell("P10", ws2)},
        "PropulsionSystem": {"mass_kg": cell("N11", ws2), "cg_m": cell("O11", ws2), "cg_pctMAC": cell("P11", ws2)},
    }
    # Batteries (Part II only)
    if ws2 is not ws:
        batt_fwd_mass = cell("N12", ws2)
        if batt_fwd_mass is not None:
            data["component_cg"]["Battery_fwd"] = {
                "mass_kg": batt_fwd_mass, "cg_m": cell("O12", ws2),
            }
        batt_aft_mass = cell("N13", ws2)
        if batt_aft_mass is not None:
            data["component_cg"]["Battery_aft"] = {
                "mass_kg": batt_aft_mass, "cg_m": cell("O13", ws2),
            }
        data["component_cg"]["CockpitSystems"] = {
            "mass_kg": cell("N14", ws2), "cg_m": cell("O14", ws2), "cg_pctMAC": cell("P14", ws2),
        }
        data["total_component_cg"] = {
            "mass_kg":   cell("N15", ws2),
            "cg_m":      cell("O15", ws2),
            "cg_pctMAC": cell("P15", ws2),
        }
    else:
        data["component_cg"]["CockpitSystems"] = {
            "mass_kg": cell("N12"), "cg_m": cell("O12"), "cg_pctMAC": cell("P12"),
        }
        data["total_component_cg"] = {
            "mass_kg":   cell("N13"),
            "cg_m":      cell("O13"),
            "cg_pctMAC": cell("P13"),
        }

    # =========================================================================
    # 10. CG EXTREMES — from Part II if available
    # =========================================================================
    data["cg_extremes"] = {
        "most_aft":    {"xcg_pctMAC": cell("AJ4", ws2), "condition": cell("AK4", ws2)},
        "most_fwd":    {"xcg_pctMAC": cell("AJ5", ws2), "condition": cell("AK5", ws2)},
        "at_OEW":      {"xcg_pctMAC": cell("AJ6", ws2), "condition": cell("AK6", ws2)},
    }

    # =========================================================================
    # 11. LOADING DIAGRAM — from Part II if available
    # =========================================================================
    ld_ws = ws2
    # Detect header row (either row 14 or row 18)
    ld_start = 15 if ws2[f"K14"].value == "Fully loaded CG" else 19
    loading = []
    for row_num in range(ld_start, ld_start + 55):
        label = ld_ws[f"K{row_num}"].value
        if label is None:
            continue
        loading.append({
            "label":         label,
            "mass_fwd":      ld_ws[f"M{row_num}"].value,
            "mass_aft":      ld_ws[f"N{row_num}"].value,
            "delta_m_fwd":   ld_ws[f"O{row_num}"].value,
            "delta_m_aft":   ld_ws[f"P{row_num}"].value,
            "cg_fwd_m":      ld_ws[f"Q{row_num}"].value,
            "cg_aft_m":      ld_ws[f"R{row_num}"].value,
            "cg_fwd_pctMAC": ld_ws[f"S{row_num}"].value,
            "cg_aft_pctMAC": ld_ws[f"T{row_num}"].value,
        })
    data["loading_diagram"] = loading

    # =========================================================================
    # 12. SEAT POSITIONS (rows 13-62)
    # =========================================================================
    seats = []
    for row_num in range(13, 63):
        pax_num = ws[f"B{row_num}"].value
        if pax_num is None or not isinstance(pax_num, (int, float)):
            continue
        seats.append({
            "pax_num":       int(pax_num),
            "pos_fwd_to_back": ws[f"C{row_num}"].value,
            "pos_back_to_fwd": ws[f"D{row_num}"].value,
        })
    data["seat_positions"] = seats

    # =========================================================================
    # 13. MAC CALCULATION (rows 105-147)
    # =========================================================================
    data["mac_calc"] = {
        "px2m":    cell("B105"),

        # Wing planform points [m] from nose
        "rootLE":   {"x": cell("F107"), "y": cell("G107")},
        "rootTE":   {"x": cell("F108"), "y": cell("G108")},
        "kinkTE":   {"x": cell("F109"), "y": cell("G109")},
        "tipLE":    {"x": cell("F110"), "y": cell("G110")},
        "tipTE":    {"x": cell("F111"), "y": cell("G111")},
        "kinkLE":   {"x": cell("F112"), "y": cell("G112")},
        "wgltLE":   {"x": cell("F113"), "y": cell("G113")},
        "wgltTE":   {"x": cell("F114"), "y": cell("G114")},

        # Panel geometry
        "panels": {
            "root_kink":      {"b_half": cell("D122"), "taper": cell("D123"),
                               "S": cell("D124"), "MAC": cell("D125"),
                               "xMAC": cell("D126"), "yMAC": cell("D127")},
            "kink_winglet":   {"b_half": cell("E122"), "taper": cell("E123"),
                               "S": cell("E124"), "MAC": cell("E125"),
                               "xMAC": cell("E126"), "yMAC": cell("E127")},
            "winglet_tip":    {"b_half": cell("F122"), "taper": cell("F123"),
                               "S": cell("F124"), "MAC": cell("F125"),
                               "xMAC": cell("F126"), "yMAC": cell("F127")},
        },

        # Overall results
        "cr":     cell("D117"),
        "MAC":    cell("D129"),
        "XLEMAC": cell("D130"),
        "YMAC":   cell("D131"),

        # Wing CG
        "wing_cg_y":   cell("D133"),
        "wing_cg_x":   cell("D141"),
        "forespar":    cell("D138"),
        "rearspar":    cell("D139"),
    }

    # =========================================================================
    # 14. AERODYNAMIC / SCISSOR-PLOT PARAMETERS (Part I only)
    # =========================================================================
    data["aero"] = {
        # Flight condition
        "M_cruise":      cell("K99"),          # cruise Mach
        "beta":          cell("K98"),          # sqrt(1 - M^2)

        # Geometry
        "sweep_c4":      cell("K106"),         # [rad] wing c/4 sweep
        "sweep_c2":      cell("N103"),         # [rad] wing c/2 sweep
        "sweep_c2_h":    cell("N99"),          # [rad] HT c/2 sweep
        "taper_ratio":   cell("K107"),         # overall wing taper ratio
        "S_net":         cell("K108"),         # [m^2] net wing area

        # Lift curve slopes
        "CLa_w":         cell("K109"),         # [1/rad] wing
        "CLa":           cell("K110"),         # [1/rad] aircraft-less-tail
        "CLa_h":         cell("K111"),         # [1/rad] horizontal tail

        # AC position components (fraction of MAC)
        "x_ac_w":        cell("K112"),         # wing AC
        "x_ac_fus":      cell("K113"),         # fuselage contribution
        "x_ac_nacelle":  cell("K114"),         # nacelle contribution
        "x_ac":          cell("K127"),         # total AC (cruise)
        "x_ac_app":      cell("K128"),         # total AC (approach)

        # Downwash
        "de_da":         cell("K116"),         # DATCOM method
        "de_da_alt":     cell("K115"),         # alternative method

        # Stability
        "SM":            cell("K129"),         # stability margin
        "l_h":           cell("K132"),         # [m] tail moment arm

        # Controllability
        "CL_h_max":      cell("K117"),         # max negative CL of HT
        "CL_Ah_max":     cell("K118"),         # max CL of aircraft-less-tail (approach)
        "Cm_ac":         cell("C164"),         # zero-lift pitching moment

        # Approach conditions
        "v_stall":       cell("K133"),         # [m/s] stall speed
        "MLW":           cell("K134"),         # [kg] max landing weight
    }

    # =========================================================================
    # 15. AERODYNAMIC / SCISSOR-PLOT PARAMETERS (Part II — CRJ-EXX)
    # =========================================================================
    if ws2 is not ws:
        data["aero_exx"] = {
            "M_cruise":      cell("K99",  ws2),
            "beta":          cell("K98",  ws2),
            "sweep_c4":      cell("K106", ws2),
            "taper_ratio":   cell("K107", ws2),
            "S_net":         cell("K108", ws2),
            "CLa_w":         cell("K109", ws2),
            "CLa":           cell("K110", ws2),
            "CLa_h":         cell("K111", ws2),
            "x_ac_w":        cell("K112", ws2),
            "x_ac_fus":      cell("K113", ws2),
            "x_ac_nacelle":  cell("K114", ws2),
            "x_ac":          cell("K127", ws2),         # x_ac_total_cruise
            "x_ac_app":      cell("K128", ws2),         # x_ac_total_approach
            "de_da":         cell("K116", ws2),
            "de_da_alt":     cell("K115", ws2),
            "SM":            cell("K129", ws2),
            "l_h":           cell("K132", ws2),         # tail moment arm
            "CL_h_max":      cell("K117", ws2),
            "CL_Ah_max":     cell("K118", ws2),
            "Cm_ac":         cell("C164", ws2),
            "v_stall":       cell("K129", ws2),
            "MLW":           cell("K130", ws2),
        }
        # CRJ-EXX CG extremes (from Part II loading diagram)
        data["cg_extremes_exx"] = {
            "most_aft":  {"xcg_pctMAC": cell("AJ4", ws2), "condition": cell("AK4", ws2)},
            "most_fwd":  {"xcg_pctMAC": cell("AJ5", ws2), "condition": cell("AK5", ws2)},
        }

    # =========================================================================
    # 16. NACELLE x_ac CONTRIBUTION (Torenbeek eq E-41)
    # d(x_ac)_nac = n * k_n * b_n^2 * l_n / (S * MAC * CLa_Ah)
    # =========================================================================
    b_n = 1.32          # [m] CF34-8C5 fan diameter
    k_n = -2.5          # rear-fuselage mounted
    n_nac = 2
    MAC_val = data["mac_calc"]["MAC"]
    S_val = data["wing"]["S"]
    XLEMAC_val = data["mac_calc"]["XLEMAC"]
    # l_n = distance from nacelle to wing c/4 (negative = behind wing)
    # Always read engine CG from Part I (unchanged per assignment spec)
    x_engine = cell("O11", ws)  # Part I propulsion system CG [m from nose]
    l_n = (XLEMAC_val + 0.25 * MAC_val) - x_engine

    CLa_Ah = data["aero"]["CLa"]
    data["nacelle"] = {
        "b_n": b_n,
        "k_n": k_n,
        "l_n": l_n,
        "x_ac_nac": n_nac * k_n * b_n**2 * l_n / (S_val * MAC_val * CLa_Ah),
    }
    # CRJ-EXX: nacelle diameter +20%
    b_n_exx = b_n * 1.20
    CLa_Ah_exx = data["aero_exx"]["CLa"] if "aero_exx" in data else CLa_Ah
    data["nacelle_exx"] = {
        "b_n": b_n_exx,
        "k_n": k_n,
        "l_n": l_n,
        "x_ac_nac": n_nac * k_n * b_n_exx**2 * l_n / (S_val * MAC_val * CLa_Ah_exx),
    }

    # =========================================================================
    # 17. DOWNWASH GRADIENT (DATCOM 1978, eq. 11.34-11.37)
    # de/da = 4.44 * [k_A * k_lam * k_H * sqrt(cos(sweep_c4))]^1.19
    #              * CLa_w(M) / CLa_w(M=0)
    # =========================================================================
    import math

    AR_val = data["wing"]["AR"]
    b_val = data["wing"]["b"]
    sweep_c4 = data["aero"]["sweep_c4"]       # [rad]
    sweep_c2 = data["aero"]["sweep_c2"]        # [rad]
    taper = data["aero"]["taper_ratio"]
    l_h_val = data["aero"]["l_h"]
    CLa_w_cruise = data["aero"]["CLa_w"]
    z_H = cell("H16")  # [m] T-tail vertical offset (from Excel)

    # k_A: aspect ratio factor (eq 11.35)
    k_A = 1.0 / AR_val - 1.0 / (1.0 + AR_val**1.7)

    # k_lam: taper factor (eq 11.36)
    k_lam = (10.0 - 3.0 * taper) / 7.0

    # k_H: tail position factor (eq 11.37)
    k_H = (1.0 - abs(z_H) / b_val) / (2.0 * l_h_val / b_val)**(1.0/3.0)

    # Mach correction: CLa_w(M) / CLa_w(M=0)
    # CLa_w at M=0 via DATCOM
    eta_datcom = 0.95
    beta_M0 = 1.0  # sqrt(1 - 0^2)
    term_M0 = (AR_val * beta_M0 / eta_datcom)**2 * (1.0 + math.tan(sweep_c2)**2 / beta_M0**2)
    CLa_w_M0 = 2.0 * math.pi * AR_val / (2.0 + math.sqrt(4.0 + term_M0))
    mach_corr = CLa_w_cruise / CLa_w_M0

    de_da_datcom = 4.44 * (k_A * k_lam * k_H * math.sqrt(math.cos(sweep_c4)))**1.19 * mach_corr

    data["downwash"] = {
        "k_A":        k_A,
        "k_lam":      k_lam,
        "k_H":        k_H,
        "z_H":        z_H,
        "mach_corr":  mach_corr,
        "CLa_w_M0":   CLa_w_M0,
        "de_da":      de_da_datcom,
    }

    # CRJ-EXX downwash (AR changes due to winglets +25%)
    if "aero_exx" in data:
        AR_exx = cell("K97", ws2)  # effective AR from Part II
        CLa_w_exx = data["aero_exx"]["CLa_w"]

        k_A_exx = 1.0 / AR_exx - 1.0 / (1.0 + AR_exx**1.7)
        term_M0_exx = (AR_exx * beta_M0 / eta_datcom)**2 * (1.0 + math.tan(sweep_c2)**2 / beta_M0**2)
        CLa_w_M0_exx = 2.0 * math.pi * AR_exx / (2.0 + math.sqrt(4.0 + term_M0_exx))
        mach_corr_exx = CLa_w_exx / CLa_w_M0_exx

        de_da_exx = 4.44 * (k_A_exx * k_lam * k_H * math.sqrt(math.cos(sweep_c4)))**1.19 * mach_corr_exx

        data["downwash_exx"] = {
            "k_A":        k_A_exx,
            "k_lam":      k_lam,
            "k_H":        k_H,
            "z_H":        z_H,
            "AR_eff":     AR_exx,
            "mach_corr":  mach_corr_exx,
            "CLa_w_M0":   CLa_w_M0_exx,
            "de_da":      de_da_exx,
        }

    # Flap / slat data
    data["flaps"] = {
        "cmacf":       cell("C162"),           # flap Cm_ac contribution
        "cmfus":       cell("C163"),           # fuselage Cm contribution
        "dCL_flap":    cell("C158"),           # delta CL_max from flaps
        "dCL_slat":    cell("C154") if cell("E154") == "DCLmax" else None,
        "dCL_HLD":     cell("C160"),           # total delta CL from HLD
    }

    wb.close()
    return data


def print_summary(data):
    """Print a formatted summary of all loaded data."""
    print("=" * 70)
    print("CRJ-1000 DATA SUMMARY")
    print("=" * 70)

    print("\n--- Weights ---")
    for k, v in data["weights"].items():
        if isinstance(v, (int, float)):
            print(f"  {k:20s} = {v:10.1f} kg")

    print("\n--- Wing ---")
    w = data["wing"]
    print(f"  Span       = {w['b']:.2f} m")
    print(f"  Area       = {w['S']:.1f} m^2")
    print(f"  AR         = {w['AR']:.3f}")

    print("\n--- Horizontal Tail ---")
    ht = data["HT"]
    print(f"  Span       = {ht['b_h']:.3f} m")
    print(f"  MAC        = {ht['c_h_mac']:.3f} m")
    print(f"  S_h        = {ht['S_h']:.3f} m^2")
    print(f"  AR_h       = {ht['AR_h']:.3f}")

    print("\n--- MAC Calculation ---")
    mc = data["mac_calc"]
    print(f"  MAC        = {mc['MAC']:.4f} m")
    print(f"  XLEMAC     = {mc['XLEMAC']:.4f} m (from nose)")

    print("\n--- Aerodynamic Parameters ---")
    a = data["aero"]
    print(f"  M_cruise   = {a['M_cruise']}")
    print(f"  CLa        = {a['CLa']:.4f} /rad")
    print(f"  CLa_h      = {a['CLa_h']:.4f} /rad")
    print(f"  de/da      = {a['de_da']:.4f}")
    print(f"  x_ac       = {a['x_ac']:.4f} MAC")
    print(f"  x_ac_app   = {a['x_ac_app']:.4f} MAC")
    print(f"  Cm_ac      = {a['Cm_ac']:.4f}")
    print(f"  CL_h_max   = {a['CL_h_max']:.4f}")
    print(f"  CL_Ah_max  = {a['CL_Ah_max']:.4f}")
    print(f"  l_h        = {a['l_h']:.3f} m")
    print(f"  SM         = {a['SM']:.2f}")

    print("\n--- CG Extremes ---")
    cge = data["cg_extremes"]
    print(f"  Most aft:     {cge['most_aft']['xcg_pctMAC']:.2f} %MAC  ({cge['most_aft']['condition']})")
    print(f"  Most forward: {cge['most_fwd']['xcg_pctMAC']:.2f} %MAC  ({cge['most_fwd']['condition']})")
    print(f"  At OEW:       {cge['at_OEW']['xcg_pctMAC']:.2f} %MAC  ({cge['at_OEW']['condition']})")


if __name__ == "__main__":
    data = load_data()
    print_summary(data)
